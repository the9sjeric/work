# 使用import导入openpyxl模块
import openpyxl
# 使用from...import从pyecharts.charts导入Timeline模块和Bar模块
from pyecharts.charts import Timeline, Bar
# 使用from...import从pyecharts导入options模块，简写为opts
from pyecharts import options as opts


# 定义函数readSheet，传入参数sheetName
# 传入需要读取的工作表，返回x轴、商品A的y轴、商品B的y轴
def readSheet(sheetName):
    # 使用wb[]读取工作表，赋值给sheet
    sheet = wb[sheetName]
    # 定义列表x_list用于存储x轴数据
    # 定义列表a_list用于存储商品A的y轴数据
    # 定义列表b_list用于存储商品B的y轴数据
    x_list = []
    a_list = []
    b_list = []
    # 使用sheet[]读取第1行数据赋值给dataX
    # 使用sheet[]读取第2行数据赋值给dataA
    # 使用sheet[]读取第3行数据赋值给dataB
    dataX = sheet[1]
    dataA = sheet[2]
    dataB = sheet[3]
    # 使用for循环配合range遍历1～7的数字
    for index in range(1, 8):
        # 使用.value配合索引获得单元格的值追击到列表x_list
        x_list.append(dataX[index].value)
        # 使用.value配合索引获得单元格的值追击到列表a_list
        a_list.append(dataA[index].value)
        # 使用.value配合索引获得单元格的值追击到列表b_list
        b_list.append(dataB[index].value)
    # 使用return以元组的方式返回x轴、商品A的y轴、商品B的y轴
    return (x_list, a_list, b_list)


# 将文件路径赋值给path
path = "024_lunbo_and_zhuzhuang.xlsx"
# 使用openpyxl.load_workbook()读取文件，赋值给wb
wb = openpyxl.load_workbook(path)

# TODO 使用Timeline创建对象赋值给tl
tl = Timeline()

# TODO 使用for循环遍历wb中的所有工作表
for sheet_name in wb.sheetnames:
    # TODO 使用Bar()函数创建对象赋值给变量bar
    bar = Bar()

    # TODO 调用readSheet函数传入sheetName，将返回值赋值给变量data
    data = readSheet(sheet_name)

    # TODO 添加参数xaxis_data=data[0]，传入add_xaxis()函数
    bar.add_xaxis(xaxis_data=data[0])

    # TODO 使用add_yaxis()函数
    # 将series_name设置商家A，data[1]赋值给y_axis
    # 将series_name设置商家B，data[2]赋值给y_axis
    bar.add_yaxis(series_name="商家A", y_axis=data[1])
    bar.add_yaxis(series_name="商家B", y_axis=data[2])

    # TODO 使用全局配置项，将标题设置为f"{sheetName}销量"
    bar.set_global_opts(title_opts=opts.TitleOpts(title=f"{sheet_name}销量"))

    # TODO 将bar赋值给chart，将sheetName赋值给time_point
    # 将参数依次传入函数tl.add()
    tl.add(chart=bar, time_point=sheet_name)

# 使用render()生成文件保存到/Users/yequ/timeline.html
tl.render("轮播柱状图.html")