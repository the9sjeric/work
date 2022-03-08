# 使用from...import从pyecharts.charts导入Funnel, Tab
from pyecharts.charts import Funnel, Tab
# 使用from...import从pyecharts导入options，简写为opts
from pyecharts import options as opts
# 使用import导入openpyxl模块
import openpyxl

# 将文件路径"020_loudoutu_fenye.xlsx"，赋值给path
path = "020_loudoutu_fenye.xlsx"
# 使用openpyxl.load_workbook()读取文件，赋值给wb
wb = openpyxl.load_workbook(path)


# 定义函数read_excel()，传入参数sheet
def read_excel(sheet):
    # 新建一个列表funnelList
    funnelList = []
    # 读取第1行内容赋值给label
    label = sheet[1]
    # 读取第2行内容赋值给num
    num = sheet[2]
    # 使用for循环遍历1～5
    for data in range(1, 6):
        # 新建列表temp
        temp = []
        # 使用if判断data等于1时
        if data == 1:
            # 使用append()将label[data]和100%追加到列表temp
            temp.append(label[data].value + "100%")
        # 否则
        else:
            # 使用(当前项/前一项)*100，赋值给pass_rate
            pass_rate = (num[data].value / num[data - 1].value) * 100
            # 使用round()保留pass_rate的一位小数，赋值给percent
            percent = round(pass_rate, 1)
            # 使用append()将标签和格式化组成的x%追加到列表temp
            temp.append(label[data].value + f"{percent}%")
        # 使用append()将num[data]追加到列表temp
        temp.append(num[data].value)
        # 使用append()将temp追加到列表funnelList
        funnelList.append(temp)
    # 使用return返回funnelList
    return funnelList


# TODO 使用Tab()函数创建实例赋值给tab
tab = Tab()
# TODO for循环遍历wb中的所有工作表的名称
for sheet_name in wb.sheetnames:
    # TODO 使用wb[]读取工作表，赋值给sheet
    sheet = wb[sheet_name]
    # TODO 调用read_excel()传入工作表名称，赋值给total_data
    total_data = read_excel(sheet)
    # TODO 使用Funnel()函数创建实例赋值给funnel
    funnel = Funnel()
    # TODO 将series_name设为空,将total_data赋值给data_pair,设置gap值为10
    # 将参数添加到add()函数中
    funnel.add(series_name="", data_pair=total_data, gap=10)

    # TODO 使用LegendOpts()，传入参数is_show=False，赋值给legend_opts，隐藏图例
    # 使用TitleOpts()，设置标题为f"{item}销售"，赋值给title_opts
    funnel.set_global_opts(title_opts=opts.TitleOpts(title=f"{sheet_name}销售"),
                           legend_opts=opts.LegendOpts(is_show=False))

    # TODO 将funnel添加到tab中，名称为工作表的名称
    tab.add(funnel, sheet_name)
# 使用render()生成漏斗图，存到路径"/Users/feifei/year.html"
tab.render("漏斗图(分页).html")