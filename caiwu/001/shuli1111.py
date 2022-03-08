import openpyxl
import pandas as pd
# 导入项目编号
project_list=["NYH202104001", "NYH202012001", "NYH201908001", "NYH202004001", "NYH202006002",
              "YC201912001", "YC202104001", "YC202010001",
              "RD202003",
              "PS1008", "PS0008",
              "GJ202001002",
              "摘要", "初期余额", "小计"]

# 导入"一次性供应商"与"线上采购"
not_project_name_list = ["一次性供应商", "线上采购"]

wb_01 = openpyxl.load_workbook("111.xlsx")
ws_01 = wb_01[wb_01.sheetnames[0]]

# 将项目有关行的R列贴上项目标签
for i in range(1,1000):
    for a in project_list:
        if a in str(ws_01[f"I{i}"].value):
            ws_01[f"R{i}"].value = a

# 将"一次性供应商"与"线上采购"的"小计"行的R列标记为"空"
for i in range(1,1000):
    if ws_01[f"H{i}"].value in not_project_name_list and ws_01[f"I{i}"].value == "小计":
        ws_01[f"R{i}"].value = None

# 选取为”平“的公司列表
ping_list = []
for i in range(1,1000):
    if ws_01[f"I{i}"].value == "小计" and ws_01[f"O{i}"].value == "平":
        ping_list.append(ws_01[f"H{i}"].value)

# 将"平"的公司的R行标记为空
for i in range(1,1000):
    if ws_01[f"H{i}"].value in ping_list:
        ws_01[f"R{i}"].value = None

name_list = []
for i in range(2,1000):
    if ws_01[f"R{i}"].value != None:
        if ws_01[f"H{i}"].value not in name_list and ws_01[f"H{i}"].value not in not_project_name_list:
            name_list.append(ws_01[f"H{i}"].value)

wb_01.save("标记平为空.xlsx")

df = pd.read_excel("标记平为空.xlsx")
df_del01 = df[df["摘要.1"].isnull()]
df.drop(index=df_del01.index,inplace=True)

df_xiaoji = df[df["摘要.1"] == "小计"]
print(df_xiaoji)
# df.drop(index=df[df["摘要.1"] == "小计"].index,inplace=True)


df.to_excel("删除了平等无效项目.xlsx")
df_xiaoji.to_excel("daicha.xlsx")

wb_xiaoji = openpyxl.load_workbook("daicha.xlsx")
ws_xiaoji = wb_xiaoji[wb_xiaoji.sheetnames[0]]
xiaoji_dir = {}
for i in range(2,500):
    xiaoji_dir[ws_xiaoji[f"I{i}"].value] = ws_xiaoji[f"R{i}"].value
print(list(xiaoji_dir.keys()))
wb_xin = openpyxl.load_workbook("删除了平等无效项目.xlsx")
ws_xin = wb_xin[wb_xin.sheetnames[0]]

for i in range(2,500):
    if ws_xin[f"I{i}"].value == None:
        break
    elif ws_xin[f"I{i}"].value in not_project_name_list:
        ws_xin[f"S{i}"].value = "OK"
    elif ws_xin[f"I{i}"].value in list(xiaoji_dir.keys()) and xiaoji_dir[ws_xin[f"I{i}"].value] == ws_xin[f"R{i}"].value :
        ws_xin[f"S{i}"].value = "OK"
    else:
        ws_xin[f"S{i}"].value = "需手工检查"
wb_xin.save("最终版.xlsx")

