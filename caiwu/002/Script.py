import openpyxl
import pandas as pd
wb01 = openpyxl.load_workbook("one.xlsx")
ws01 = wb01[wb01.sheetnames[0]]

wb02 = openpyxl.load_workbook("two.xlsx")
ws02 = wb02[wb02.sheetnames[0]]

ws02.insert_cols(5)
ws02.insert_cols(5)
ws02["E1"].value = "类别"
ws02["F1"].value = "项目编号"

project_dic = {}
for num in range(2,1000):
    if ws01[f"E{num}"].value == "初期余额":
        continue
    elif ws01[f"E{num}"].value == None:
        break
    elif ws01[f"D{num}"].value != None:
        project_dic[ws01[f"D{num}"].value] = ws01[f"J{num}"].value

for num in range(2,1000):
    if ws02[f"L{num}"].value == None:
        break
    elif ws02[f"I{num}"].value == "贷" or ws02[f"G{num}"].value == "合计":
        ws02[f"A{num}"].value = None
        ws02[f"B{num}"].value = None
        ws02[f"C{num}"].value = None
        ws02[f"D{num}"].value = None
        ws02[f"E{num}"].value = None
        ws02[f"F{num}"].value = None
        ws02[f"G{num}"].value = None
        ws02[f"H{num}"].value = None
        ws02[f"I{num}"].value = None
        ws02[f"J{num}"].value = None
        ws02[f"K{num}"].value = None
        ws02[f"L{num}"].value = None

chai_lv_fei = ["差旅费", "市内交通费"]
ye_wu_fei = ["业务费"]
ren_gong_fei = ["工资奖金", "职工社保", "职工公积金", "职工福利费"]
for num in range(2,1000):
    if ws02[f"D{num}"].value == None:
        break
    elif ws02[f"D{num}"].value in chai_lv_fei:
        ws02[f"E{num}"].value = "差旅费"
    elif ws02[f"D{num}"].value in ye_wu_fei:
        ws02[f"E{num}"].value = "业务费"
    elif ws02[f"D{num}"].value in ren_gong_fei:
        ws02[f"E{num}"].value = "人工费"
    else:
        ws02[f"E{num}"].value = "其他"

for num in range(2,1000):
    if ws02[f"B{num}"].value == None:
        break
    else:
        ws02[f"F{num}"].value = project_dic[ws02[f"B{num}"].value]

wb02.save("two_new.xlsx")

