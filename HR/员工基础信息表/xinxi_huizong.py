import time
import openpyxl
import os
import datetime

print("脚本即将开始运行。")
time.sleep(2)
print("脚本已经启动，请稍等。")
begin_time = time.time()
"""excel数字时间转换字符串函数"""
def timetostr(figure_time):
    shi = openpyxl.utils.datetime.from_excel(
        figure_time, epoch=datetime.datetime(1899, 12, 30, 0, 0), timedelta=False)
    jian = shi.strftime("%Y年%#m月")
    return jian

"""获取所有文件名"""
all_items = os.listdir('./员工基础信息')
xlsx_items = []
for item in all_items:
    if '.xlsx' in item:
        xlsx_items.append(item)

mubanname = "员工基础信息汇总表（模板）.xlsx"

wb_muban = openpyxl.load_workbook(mubanname)
ws_muban = wb_muban[wb_muban.sheetnames[0]]

for filename in xlsx_items:
    path = f"./员工基础信息/{filename}"
    wb = openpyxl.load_workbook(path)
    ws = wb[wb.sheetnames[0]]

    xuhao = ""
    xingming = ws['B3'].value
    gangwei = ws['D5'].value
    zhiji = ws['B5'].value

    if type(ws['H4'].value) == int:
        rusishijian = timetostr(ws['H4'].value)
    else:
        rusishijian = ws['H4'].value

    gongzuojingyan = ws['J4'].value

    if type(ws['A9'].value) == int:
        starttime = timetostr(ws['A9'].value)
    else:
        starttime = ws['A9'].value

    if type(ws['B9'].value) == int:
        endtime = timetostr(ws['B9'].value)
    else:
        endtime = ws['B9'].value

    qiangongsi = ws['C9'].value
    qiangangwei = ws['E9'].value

    jiangcheng = f"{ws['A19'].value} {str(ws['B19'].value)} {str(ws['C19'].value)} {str(ws['D19'].value)}\n" \
                 f"{ws['A20'].value} {str(ws['B20'].value)} {str(ws['C20'].value)} {str(ws['D20'].value)}\n" \
                 f"{ws['A21'].value} {str(ws['B21'].value)} {str(ws['C21'].value)} {str(ws['D21'].value)}"

    pinggushijian = ""
    nengli = ""
    taidu = ""

    guihua = ws['B24'].value
    jihua = ws['B25'].value
    xiangmu = f"{ws['B26'].value}\n" \
              f"{ws['E26'].value}\n" \
              f"{ws['H26'].value}"
    zhuanli = ws['B27'].value
    data_list = []
    data_list.extend([xuhao, xingming, gangwei, zhiji, rusishijian, gongzuojingyan, starttime, endtime,
                     qiangongsi, qiangangwei, jiangcheng, pinggushijian, nengli, taidu, guihua,
                     jihua, xiangmu, zhuanli])
    ws_muban.append(data_list)

wb_muban.save("666.xlsx")

end_time = time.time()
run_time = end_time - begin_time
print(f"脚本已经运行顺利完成。\n本次共用时{run_time}秒。\n本窗口将在2秒钟以后自动关闭，谢谢使用。")
time.sleep(2)