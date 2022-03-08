import openpyxl
import os
'''历遍原始数据'''
all_items = os.listdir('./data')
xlsx_item = []
for item in all_items:
    if '.xlsx' in item:
        xlsx_item.append(item)

'''建立汇总表格以及表头'''
new_wb = openpyxl.Workbook()
new_ws = new_wb['Sheet']
new_ws['A1'].value = '项目名称'
new_ws['B1'].value = '项目经理'
new_ws['C1'].value = '检查阶段'
new_ws['D1'].value = '×的数量'
new_ws['E1'].value = '√的数量'
new_ws['F1'].value = '不适用的数量'
new_ws['G1'].value = '审计符合率'

for data in xlsx_item:
    wb = openpyxl.load_workbook(f'data/{data}', data_only=True)
    sheet_num = len(wb.sheetnames)
    for num in range(1, sheet_num):
        ws = wb[wb.sheetnames[num]]
        q_list = []
        try:
            q_list.append(ws[2][2].value)
        except IndexError as e:
            print(f'{data}\n'
                  f'{wb.sheetnames[num]}\n'
                  f'该页面格式不对,没有进行数据录入!!!\n'
                  f'------------------------------')
            continue

        try:
            q_list.append(ws[2][4].value)
        except IndexError as e:
            print(f'{data}\n'
                  f'{wb.sheetnames[num]}\n'
                  f'该页面格式不对,没有进行数据录入!!!\n'
                  f'-------------------------------')
            continue

        q_list.append(ws.title)
        x_num = 0
        y_num = 0
        z_num = 0
        for cell in ws['E']:
            if cell.value == '√':
                x_num = x_num + 1
            elif cell.value == '×':
                y_num = y_num + 1
            elif cell.value == '不适用':
                z_num = z_num + 1
        q_list.append(x_num)
        q_list.append(y_num)
        q_list.append(z_num)
        try:
            ok = round(x_num / (x_num + y_num) * 100, 2)
        except ZeroDivisionError as e :
            print(f'{data}\n'
                  f'{wb.sheetnames[num]}\n'
                  f'审计符合率计算报错,请检查。本页没有进行数据录入!!!\n'
                  f'==========================================')
            continue
        q_list.append(ok)
        new_ws.append(q_list)

new_wb.save('666.xlsx')