import openpyxl

source_file = "ipguard_source/ip-guard推送清单.xlsx"
target_file = "ipguard_source/杭州电脑_ipguard安装清单(固定资产版).xlsx"

wb_source = openpyxl.load_workbook(source_file)
wb_target = openpyxl.load_workbook(target_file)

ws_source = wb_source[wb_source.sheetnames[0]]
ws_target = wb_target[wb_target.sheetnames[0]]

source_dic = {}
for number in range(1, 200):
    if ws_source[f"C{number}"].value != None:
        source_dic[ws_source[f"C{number}"].value] = \
            [ws_source[f"A{number}"].value, ws_source[f"B{number}"].value, ws_source[f"D{number}"].value]

for row in ws_target.rows:
    if row[6].value in list(source_dic.keys()):
        row[4].value = source_dic[row[6].value][0]
        row[5].value = source_dic[row[6].value][1]
        row[7].value = source_dic[row[6].value][2]

wb_target.save("目标文件.xlsx")
