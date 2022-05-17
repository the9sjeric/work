import openpyxl
import pandas as pd
import datetime

"""定义数据函数名"""
zhiji_file = "source/员工岗位职级.xlsx"
mingxi_file = "source/工时明细.xlsx"
caigou_file = "source/采购、差旅费-模板.xlsx"
jiancha_file = "中间数据表(检查用).xlsx"
target_file = "source/目标文件.xlsx"
target_new_file = "目标文件(新).xlsx"

"""excel数字时间转换字符串函数"""
def timetostr(figure_time):
    shi = openpyxl.utils.datetime.from_excel(
        figure_time, epoch=datetime.datetime(1899, 12, 30, 0, 0), timedelta=False)
    jian = shi.strftime("%Y年%#m月")
    return jian

"""制作<人时hash表>"""
zhiji_hash = {}
zhiji_df = pd.read_excel(zhiji_file)
for money in range(0, len(zhiji_df)):
    zhiji_hash[zhiji_df.loc[money]["姓名"]] = zhiji_df.loc[money]["人时"]

"""制作<总费用表>"""
rengongfei_df = pd.read_excel(mingxi_file)
df = rengongfei_df[['填报日期', '员工姓名', '折算时间', '项目编号']].copy()
df["填报日期"] = pd.to_datetime(df["填报日期"])
df["填报日期"] = df["填报日期"].dt.strftime("%Y年%#m月")
df['人时'] = 0

for i in range(0, len(df)):
    if df.loc[i, "员工姓名"] in zhiji_hash.keys():
        df.loc[i, "人时"] = zhiji_hash[df.loc[i, '员工姓名']]
df["人工费"] = df["折算时间"] * df["人时"]
df_group = df.groupby("项目编号").sum()
df_group.to_excel(jiancha_file)

"""提取 <时间> <项目编号> <折算时间> 和 <人工>"""
shijian = df.loc[0, "填报日期"]

pro_cost = {} #创建项目与时间人工的字典
for i in range(0, len(df_group)):
    sjrg = []    #创建时间和人工list
    sjrg.append(df_group.iloc[i]["折算时间"].round(2))
    sjrg.append(df_group.iloc[i]["人工费"].round(2))
    pro_cost[list(df_group.index)[i]] = sjrg

"""时间人工写入目标文件"""
wb = openpyxl.load_workbook(target_file)

for i in wb.sheetnames:
    ws = wb[i]
    if i in list(pro_cost.keys()):  #如果项目名在费用列表中则，则填入相应数字
        for number in range(1, 100):
            if  type(ws[f"A{number}"].value) == int and timetostr(ws[f"A{number}"].value) == shijian:
                ws[f"B{number}"].value = pro_cost[i][0]
                ws[f"C{number}"].value = pro_cost[i][1]
    else:                           #如果项目名没有在列表中，则不填
        for number in range(1, 100):
            if type(ws[f"A{number}"].value) == int and timetostr(ws[f"A{number}"].value) == shijian:
                ws[f"B{number}"].value = None
                ws[f"C{number}"].value = None

"""获取采购、差旅、外包、其他4种数据"""
wb_caigou = openpyxl.load_workbook(caigou_file, data_only=True)
caigou_dic = {}
for sheetname in wb_caigou.sheetnames:
    ws = wb_caigou[sheetname]
    data_list = {}
    for number in range(1, 100):   #先获取时间数据，以便后面进行判断
        if type(ws[f"A{number}"].value) == int:
            data_list[ws[f"A{number}"].value] = [ws[f"D{number}"].value, ws[f"E{number}"].value,
                                                 ws[f"F{number}"].value, ws[f"G{number}"].value]
        elif ws[f"A{number}"].value == "合计":
            break
    caigou_dic[sheetname] = data_list
"""对目标文件写入采购、差旅、外包、其他4种数据"""
for q in wb.sheetnames:
    ws = wb[q]
    if q in wb_caigou.sheetnames:
        for number in range(1, 100):
            if type(ws[f"A{number}"].value) == int and ws[f"A{number}"].value in caigou_dic[q].keys():
                print(caigou_dic[q].keys())
                print(number)
                print(ws[f"A{number}"].value)
                ws[f"D{number}"].value = caigou_dic[q][ws[f"A{number}"].value][0]
                ws[f"E{number}"].value = caigou_dic[q][ws[f"A{number}"].value][1]
                ws[f"F{number}"].value = caigou_dic[q][ws[f"A{number}"].value][2]
                ws[f"G{number}"].value = caigou_dic[q][ws[f"A{number}"].value][3]
    else:
        continue

"""将财务数据的图拷贝至目标文件中"""
for name in wb_caigou.sheetnames:
    if name in wb.sheetnames:
        for image in wb_caigou[name]._images:
            wb[name].add_image(image)
            #print(image.anchor._from.row)

wb.save(target_new_file)