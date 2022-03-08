import openpyxl
import pandas as pd
import datetime

"""定义数据函数名"""
zhiji_file = "员工岗位职级.xlsx"
mingxi_file = "工时明细.xlsx"
caigou_file = "采购、差旅费-模板.xlsx"
jiancha_file = "中间数据表(检查用).xlsx"
target_file = "目标文件.xlsx"
target_new_file = "目标文件(新).xlsx"

"""excel数字时间转换字符串函数"""
def timetostr(figure_time):
    shi = openpyxl.utils.datetime.from_excel(
        figure_time, epoch=datetime.datetime(1899, 12, 30, 0, 0), timedelta=False)
    jian = shi.strftime("%Y年%#m月")
    return jian

"""制作<人时hash表>"""
zhiji_hash = {}
zhiji_df = pd.read_excel(f"source/{zhiji_file}")
for money in range(0, len(zhiji_df)):
    zhiji_hash[zhiji_df.loc[money]["姓名"]] = zhiji_df.loc[money]["人时"]

"""制作<总费用表>"""
rengongfei_df = pd.read_excel(f"source/{mingxi_file}")
df = rengongfei_df[['填报日期', '员工姓名', '折算时间', '项目编号']].copy()
df["填报日期"] = pd.to_datetime(df["填报日期"])
df["填报日期"] = df["填报日期"].dt.strftime("%Y年%#m月")
df['人时'] = 0

for i in range(0, len(df)):
    if df.loc[i, "员工姓名"] in zhiji_hash.keys():
        df.loc[i, "人时"] = zhiji_hash[df.loc[i, '员工姓名']]
df["人工费"] = df["折算时间"] * df["人时"]
df_group = df.groupby("项目编号").sum()
df_group.to_excel(jiancha_file)

"""提取 <时间> <项目编号> <折算时间> 和 <人工>"""
shijian = df.loc[0, "填报日期"]

pro_cost = {} #创建项目与时间人工的字典
for i in range(0, len(df_group)):
    sjrg = []    #创建时间和人工list
    sjrg.append(df_group.iloc[i]["折算时间"].round(2))
    sjrg.append(df_group.iloc[i]["人工费"].round(2))
    pro_cost[list(df_group.index)[i]] = sjrg

"""时间人工写入目标文件"""
wb = openpyxl.load_workbook(f"source/{target_file}")

for i in wb.sheetnames:
    ws = wb[i]
    if i in list(pro_cost.keys()):  #如果项目名在费用列表中则，则填入相应数字
        for number in range(1, 100):
            if  type(ws[f"A{number}"].value) == int and timetostr(ws[f"A{number}"].value) == shijian:
                ws[f"B{number}"].value = pro_cost[i][0]
                ws[f"C{number}"].value = pro_cost[i][1]
    else:                           #如果项目名没有在列表中，则不填
        for number in range(1, 100):
            if type(ws[f"A{number}"].value) == int and timetostr(ws[f"A{number}"].value) == shijian:
                ws[f"B{number}"].value = None
                ws[f"C{number}"].value = None

"""获取采购数据"""
wb_caigou = openpyxl.load_workbook(f"source/{caigou_file}")
caigou_dic = {}
for sheetname in wb_caigou.sheetnames:
    ws = wb_caigou[sheetname]
    data_list = []
    for number in range(1, 100):   #先获取时间数据，以便后面进行判断
        if type(ws[f"A{number}"].value) == int:
            data_list.append(timetostr(ws[f"A{number}"].value))
        else:
            data_list.append(ws[f"A{number}"].value)
    cg_list = []

    hangshu = 0
    if shijian not in data_list:    #判断是否有本月数据，“没有”则自动填写为0
        cg_list.extend([None, None, None, None])
        caigou_dic[sheetname] = cg_list
    else:                           #判断是否有本月数据，”有“则正常填写
        for number in range(1, 100):        #定位本月数据所在行
            if type(ws[f"A{number}"].value) == int and timetostr(ws[f"A{number}"].value) == shijian:
                hangshu = number
            elif ws[f"A{number}"].value == "期间":
                title_hangshu = number
        if ws[f"F{title_hangshu}"].value == "外包费" and ws[f"G{title_hangshu}"].value == "其他":
            cg_list.extend([ws[f"D{hangshu}"].value,ws[f"E{hangshu}"].value,
                            ws[f"F{hangshu}"].value,ws[f"G{hangshu}"].value])
        elif ws[f"F{title_hangshu}"].value == "外包费" and ws[f"G{title_hangshu}"].value != "其他":
            cg_list.extend([ws[f"D{hangshu}"].value, ws[f"E{hangshu}"].value,
                            ws[f"F{hangshu}"].value, None])
        elif ws[f"F{title_hangshu}"].value == "其他":
            cg_list.extend([ws[f"D{hangshu}"].value, ws[f"E{hangshu}"].value,
                            None, ws[f"G{hangshu}"].value])
        else:
            cg_list.extend([ws[f"D{hangshu}"].value, ws[f"E{hangshu}"].value,
                            None, None])
    caigou_dic[sheetname] = cg_list
print(caigou_dic)
"""对目标文件写入采购数据"""
for q in wb.sheetnames:
    ws = wb[q]
    hangshu = 0
    for number in range(1, 100):     #定位本月数据所在行
        if type(ws[f"A{number}"].value) == int and timetostr(ws[f"A{number}"].value) == shijian:
            hangshu = number
    ws[f"D{hangshu}"].value = caigou_dic[q][0]
    ws[f"E{hangshu}"].value = caigou_dic[q][1]
    ws[f"F{hangshu}"].value = caigou_dic[q][2]
    ws[f"G{hangshu}"].value = caigou_dic[q][3]

wb.save(f"{target_new_file}")