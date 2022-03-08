import openpyxl

wb = openpyxl.load_workbook(end_file)
wb.create_sheet("数据透视")

wb_end = openpyxl.load_workbook("数据透视表.xlsx")
sheet = wb_end[wb_end.sheetnames[0]]
for value in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column, values_only=True):
    value = list(value)
    wb["数据透视"].append(value)

wb.save(end_file)